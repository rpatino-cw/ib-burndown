#!/usr/bin/env python3
"""EVI01 IB Lookup — search switch, see connections + map + tips."""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import webbrowser
from datetime import datetime

from rich.console import Console
import openpyxl

try:
    from cwhelper.config import BOLD, DIM, RESET, RED, GREEN, YELLOW, CYAN, WHITE, MAGENTA, BLUE
except ImportError:
    BOLD, DIM, RESET = "\033[1m", "\033[2m", "\033[0m"
    RED, GREEN, YELLOW = "\033[31m", "\033[32m", "\033[33m"
    CYAN, WHITE, MAGENTA, BLUE = "\033[36m", "\033[97m", "\033[35m", "\033[34m"

_DIR = os.path.dirname(os.path.abspath(__file__))
_SKETCH_XLSX = os.path.join(_DIR, "EVI01 - IB Sketch.xlsx")
_CACHE_PATH = os.path.join(_DIR, ".ib_lookup_cache.json")

console = Console(highlight=False)

# ════════════════════════════════════════════════════════════════════
#  TIPS — hands-on DCT level (no SSH/Teleport)
# ════════════════════════════════════════════════════════════════════

_TIPS = [
    ("Find the racks",       "Go to ~R{rack_a} ({src_name}) and ~R{rack_b} ({dest_name})"),
    ("Check cable seating",  "Reseat cable at both ends — pull fully out, inspect, push back in"),
    ("Inspect for damage",   "Look for bent pins, dust, kinks, or damaged fiber ends"),
    ("Verify port LEDs",     "Check link LED on both switch ports — should be green/amber, not off"),
    ("Label check",          "Confirm cable label matches cutsheet for both ends"),
    ("Escalate if no fix",   "If reseat doesn't fix it, comment on the Google Sheet and tag network"),
]


def _get_tips(conn: dict) -> list[tuple[str, str]]:
    rack_a = _extract_rack(conn.get("src_name", ""), conn.get("data_hall", ""))
    rack_b = _extract_rack(conn.get("dest_name", ""), conn.get("data_hall", ""))
    subs = {
        "src_name": conn.get("src_name", ""),
        "dest_name": conn.get("dest_name", ""),
        "src_port": conn.get("src_port", ""),
        "dest_port": conn.get("dest_port", ""),
        "rack_a": rack_a or "?",
        "rack_b": rack_b or "?",
    }
    filled = []
    for label, cmd in _TIPS:
        try:
            filled.append((label, cmd.format(**subs)))
        except KeyError:
            filled.append((label, cmd))
    return filled


# ════════════════════════════════════════════════════════════════════
#  NORMALIZATION (borrowed from cwhelper/services/ib_trace.py)
# ════════════════════════════════════════════════════════════════════

_TYPE_PREFIX = {"Spine": "S", "Core": "C", "Leaf": "L", "Node": "N"}


def _normalize_port(value) -> str:
    """Convert datetime-corrupted port values back to 'port/lane'."""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, (int, float)):
        i = int(value)
        return str(i) if value == i else str(value)
    if value is None:
        return ""
    return str(value).strip()


def _normalize_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _normalize_cab(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        i = int(value)
        return str(i) if value == i else str(value)
    return str(value).strip()


def _build_switch_name(type_str: str, switch_id: str, data_hall: str, cab: str = "") -> str:
    """Build canonical switch name: S1.1.1, L10.1.2-DH2, C1.4."""
    prefix = _TYPE_PREFIX.get(type_str, "")
    if type_str == "Leaf" and cab:
        name = f"{prefix}{cab}.{switch_id}"
        if data_hall:
            name += f"-{data_hall}"
    else:
        name = f"{prefix}{switch_id}"
    return name


# ════════════════════════════════════════════════════════════════════
#  DATA — IB Sketch pull schedule tabs (primary data source)
# ════════════════════════════════════════════════════════════════════

_PREFIX_TYPE = {"S": "Spine", "C": "Core", "L": "Leaf", "N": "Node"}


def _parse_switch_name(name: str) -> dict:
    """Parse a switch name like 'C1.17', 'S1.5.2', 'L10.1.2-DH2' into components."""
    result = {"type": "", "dh": "", "cab": "", "id": ""}
    if not name:
        return result

    # Extract DH suffix
    dh_m = re.search(r'-DH(\d+)$', name, re.IGNORECASE)
    base = name[:dh_m.start()] if dh_m else name
    result["dh"] = f"DH{dh_m.group(1)}" if dh_m else ""

    # Extract prefix letter
    if base and base[0].upper() in _PREFIX_TYPE:
        result["type"] = _PREFIX_TYPE[base[0].upper()]
        base = base[1:]
    else:
        return result

    # For Leaf: first number is cab, rest is id (e.g. L10.1.2 → cab=10, id=1.2)
    if result["type"] == "Leaf":
        m = re.match(r'^(\d+)\.(.+)$', base)
        if m:
            result["cab"] = m.group(1)
            result["id"] = m.group(2)
        else:
            result["id"] = base
    else:
        # Core/Spine: first number group could include cab context
        # C1.17 → group 1, switch 17; S1.5.2 → group 1, tier 5, pos 2
        result["id"] = base

    return result


def _parse_connections_from_sketch(path: str) -> tuple[list[dict], dict[str, dict]]:
    """Parse Pull Schedule tabs from IB Sketch into connections + enrichment lookup.

    Returns (connections, enrichment_dict).
    """
    if not os.path.isfile(path):
        print(f"  {RED}Not found:{RESET} {path}")
        return [], {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    connections: list[dict] = []
    enrichment: dict[str, dict] = {}

    for tab_name in wb.sheetnames:
        if "Pull Schedule" not in tab_name:
            continue

        # Infer DH from tab name
        tab_dh = ""
        if "DH1" in tab_name:
            tab_dh = "DH1"
        elif "DH2" in tab_name:
            tab_dh = "DH2"
        elif "DH3" in tab_name:
            tab_dh = "DH3"

        ws = wb[tab_name]
        header_row = None
        col_map = {}

        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header_row = row
            break

        if not header_row:
            continue

        for i, h in enumerate(header_row):
            if h is None:
                continue
            hl = str(h).strip().lower()
            if hl == "status":
                col_map["status"] = i
            elif hl == "source":
                col_map["source"] = i
            elif hl == "source port":
                col_map["src_port"] = i
            elif hl == "destination":
                col_map["dest"] = i
            elif hl == "destination port":
                col_map["dest_port"] = i
            elif hl == "cable type":
                col_map["cable_type"] = i
            elif hl == "cable length":
                col_map["cable_length"] = i
            elif "optic" in hl and "type" in hl:
                col_map["optic_type"] = i
            elif hl == "fabric id":
                col_map["fabric_id"] = i

        if "source" not in col_map or "dest" not in col_map:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            def _g(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                val = row[idx]
                if isinstance(val, datetime):
                    return f"{val.month}/{val.day}"
                return str(val).strip()

            src_name = _g("source")
            dest_name = _g("dest")
            if not src_name or not dest_name:
                continue

            src_port = _g("src_port")
            dest_port = _g("dest_port")
            status = _g("status")
            cable_type = _g("cable_type")
            cable_length = _g("cable_length")
            optic_type = _g("optic_type")
            fabric_id = _g("fabric_id")

            # Parse switch name components
            src_info = _parse_switch_name(src_name)
            dest_info = _parse_switch_name(dest_name)

            # Determine data hall
            data_hall = src_info["dh"] or dest_info["dh"] or tab_dh

            connections.append({
                "data_hall": data_hall,
                "src_type": src_info["type"],
                "src_dh": src_info["dh"] or tab_dh,
                "src_cab": src_info["cab"],
                "src_id": src_info["id"],
                "src_port": src_port,
                "src_name": src_name,
                "dest_type": dest_info["type"],
                "dest_dh": dest_info["dh"] or tab_dh,
                "dest_cab": dest_info["cab"],
                "dest_id": dest_info["id"],
                "dest_port": dest_port,
                "dest_name": dest_name,
                "tab_ref": tab_name,
                "status": status,
                "cable_type": cable_type,
                "cable_length": cable_length,
                "optic_type": optic_type,
                "fabric_id": fabric_id,
            })

            # Also build enrichment lookup for backwards compat
            key = f"{src_name.upper()}|{src_port}|{dest_name.upper()}|{dest_port}"
            enrichment[key] = {
                "status": status,
                "cable_type": cable_type,
                "cable_length": cable_length,
                "optic_type": optic_type,
                "fabric_id": fabric_id,
            }

    wb.close()
    return connections, enrichment


def _parse_elevations(path: str) -> dict[str, dict]:
    """Parse ELEV tabs. Returns lookup: switch_name → {rack, ru, sku, dh, row}."""
    if not os.path.isfile(path):
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    elevations: dict[str, dict] = {}

    for tab_name in wb.sheetnames:
        if "ELEV" not in tab_name:
            continue

        ws = wb[tab_name]
        rows_data = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows_data.append(row)
            if len(rows_data) >= 55:
                break

        if len(rows_data) < 4:
            continue

        # Row 2 (idx 1): DH + Row label
        row2 = rows_data[1]
        row_label = str(row2[2]).strip() if row2[2] else ""
        # Extract DH from tab name
        dh = ""
        if "DH1" in tab_name:
            dh = "DH1"
        elif "DH2" in tab_name:
            dh = "DH2"
        elif "DH3" in tab_name:
            dh = "DH3"

        # Row 3 (idx 2): rack numbers at cols 3, 5, 7, 9, ...
        row3 = rows_data[2]
        rack_cols = {}  # col_index → rack_number
        for ci in range(3, len(row3), 2):
            val = row3[ci]
            if val is not None:
                try:
                    rack_cols[ci] = int(float(str(val)))
                except (ValueError, TypeError):
                    pass

        # Data rows (idx 3+): col 0 = RU, rack cols = switch names, rack cols+1 = SKU
        for row in rows_data[3:]:
            if not row or row[0] is None:
                continue
            try:
                ru = int(float(str(row[0])))
            except (ValueError, TypeError):
                continue

            for ci, rack_num in rack_cols.items():
                if ci >= len(row) or row[ci] is None:
                    continue
                switch_name = str(row[ci]).strip()
                if not switch_name or switch_name in ("Labels", "POWER", "Rack#"):
                    continue

                sku = ""
                if ci + 1 < len(row) and row[ci + 1] is not None:
                    sku = str(row[ci + 1]).strip()
                    if sku == "CW-SKU":
                        sku = ""

                elevations[switch_name.upper()] = {
                    "rack": rack_num,
                    "ru": ru,
                    "sku": sku,
                    "dh": dh,
                    "row": row_label,
                }

    wb.close()
    return elevations


# ════════════════════════════════════════════════════════════════════
#  CACHING
# ════════════════════════════════════════════════════════════════════

def _load_cache() -> dict | None:
    if not os.path.isfile(_CACHE_PATH):
        return None
    try:
        sketch_mtime = os.path.getmtime(_SKETCH_XLSX) if os.path.isfile(_SKETCH_XLSX) else 0
        with open(_CACHE_PATH) as f:
            cache = json.load(f)
        if cache.get("sketch_mtime") == sketch_mtime and "elevations" in cache:
            return {"connections": cache["connections"], "elevations": cache["elevations"]}
    except (json.JSONDecodeError, OSError, KeyError):
        pass
    return None


def _save_cache(connections: list[dict], elevations: dict[str, dict]):
    try:
        sketch_mtime = os.path.getmtime(_SKETCH_XLSX) if os.path.isfile(_SKETCH_XLSX) else 0
        with open(_CACHE_PATH, "w") as f:
            json.dump({
                "sketch_mtime": sketch_mtime,
                "connections": connections,
                "elevations": elevations,
            }, f)
    except OSError:
        pass


def _load_data() -> tuple[list[dict], dict[str, dict]]:
    cached = _load_cache()
    if cached is not None:
        return cached["connections"], cached["elevations"]

    connections, _ = _parse_connections_from_sketch(_SKETCH_XLSX)
    elevations = _parse_elevations(_SKETCH_XLSX)
    if connections:
        _save_cache(connections, elevations)
    return connections, elevations


# ════════════════════════════════════════════════════════════════════
#  RACK EXTRACTION
# ════════════════════════════════════════════════════════════════════

def _extract_rack(switch_name: str, data_hall: str = "") -> int | None:
    """Get rack number for a switch. Uses elevation data if available, falls back to cab number."""
    if not switch_name:
        return None
    # Use real elevation data first (most accurate)
    elev = _ELEVATIONS.get(switch_name.upper())
    if elev:
        return elev["rack"]
    # Fallback: Leaf cab number from name
    m = re.match(r'^L(\d+)\.', switch_name)
    if m:
        return int(m[1])
    return None


# ════════════════════════════════════════════════════════════════════
#  LAYOUT
# ════════════════════════════════════════════════════════════════════

def _load_layout(hall_key: str = "US-CENTRAL-07A.DH2"):
    paths = [
        os.path.join(os.path.expanduser("~/dev/cw-node-helper"), "dh_layouts.json"),
        os.path.join(_DIR, "dh_layouts.json"),
    ]
    for p in paths:
        try:
            with open(p) as f:
                data = json.load(f)
            layout = data.get(hall_key)
            if layout:
                rpr = layout.get("racks_per_row", 10)
                return {
                    "label": hall_key.split(".")[-1],
                    "columns": [
                        {"label": c["label"], "start": c["start"],
                         "num_rows": c["num_rows"], "rpr": c.get("racks_per_row", rpr)}
                        for c in layout["columns"]
                    ],
                    "serpentine": layout.get("serpentine", True),
                }
        except (FileNotFoundError, json.JSONDecodeError, KeyError):
            continue

    # Fallback for DH2
    if "DH2" in hall_key:
        return {
            "label": "DH2",
            "columns": [
                {"label": "Left", "start": 1, "num_rows": 12, "rpr": 10},
                {"label": "Right", "start": 121, "num_rows": 14, "rpr": 10},
            ],
            "serpentine": True,
        }
    # Fallback for DH1
    return {
        "label": "DH1",
        "columns": [
            {"label": "Left", "start": 1, "num_rows": 14, "rpr": 10},
            {"label": "Right", "start": 141, "num_rows": 17, "rpr": 10},
        ],
        "serpentine": True,
    }


_LAYOUTS = {
    "DH1": _load_layout("US-CENTRAL-07A.DH1"),
    "DH2": _load_layout("US-CENTRAL-07A.DH2"),
}

# Global elevation lookup — populated at runtime by _load_data()
_ELEVATIONS: dict[str, dict] = {}


# ════════════════════════════════════════════════════════════════════
#  MAP
# ════════════════════════════════════════════════════════════════════

def _rack_at(col: dict, row: int, pos: int, serpentine: bool) -> int:
    rpr = col["rpr"]
    if serpentine and row % 2 == 1:
        return col["start"] + (row + 1) * rpr - 1 - pos
    return col["start"] + row * rpr + pos


def _col_width(rpr):
    return 4 + rpr * 2 - 1 + 1 + 4


def _draw_map(layout: dict,
              highlight_a: int | None = None,
              highlight_b: int | None = None,
              label_a: str = "",
              label_b: str = ""):
    serp = layout["serpentine"]
    cols = layout["columns"]
    COL_GAP = "       "

    legend = []
    if highlight_a and label_a:
        legend.append(f"  {CYAN}{BOLD}@{RESET} = {label_a} {DIM}(~R{highlight_a}){RESET}")
    if highlight_b and label_b:
        legend.append(f"  {YELLOW}{BOLD}#{RESET} = {label_b} {DIM}(~R{highlight_b}){RESET}")

    print(f"\n  {BOLD}{layout['label']} Floor Map{RESET}")
    for l in legend:
        print(l)
    print()

    hdr_parts = []
    for col in cols:
        rpr = col["rpr"]
        end_rack = col["start"] + col["num_rows"] * rpr - 1
        part = f"{col['label']} (R{col['start']}–R{end_rack})"
        cw = _col_width(rpr)
        hdr_parts.append(f"{part:<{cw}}")
    print(f"  {DIM}{COL_GAP.join(hdr_parts)}{RESET}")

    max_rows = max(col["num_rows"] for col in cols)

    def _cell(rn):
        if rn == highlight_a:
            return f"{CYAN}{BOLD}@{RESET}"
        if rn == highlight_b:
            return f"{YELLOW}{BOLD}#{RESET}"
        return f"{DIM}-{RESET}"

    for row in range(max_rows):
        col_strs = []
        for ci, col in enumerate(cols):
            rpr = col["rpr"]
            cw = _col_width(rpr)
            if row < col["num_rows"]:
                first_rack = _rack_at(col, row, 0, serp)
                last_rack = _rack_at(col, row, rpr - 1, serp)
                label_l = f"{DIM}{first_rack:>3}{RESET} "
                label_r = f" {DIM}{last_rack:<3}{RESET}"
                cells = " ".join(_cell(_rack_at(col, row, pos, serp)) for pos in range(rpr))
                col_strs.append(label_l + cells + label_r)
            else:
                col_strs.append(" " * cw)

        print(f"  {col_strs[0]}{COL_GAP}{col_strs[1] if len(col_strs) > 1 else ''}")

        if row % 2 == 1 and row < max_rows - 1:
            print()
    print()


# ════════════════════════════════════════════════════════════════════
#  SEARCH
# ════════════════════════════════════════════════════════════════════

def _switch_matches(name: str, pattern: str) -> bool:
    name_up = name.upper()
    pat_up = pattern.upper()
    if name_up == pat_up:
        return True
    if "-" in name_up and "-" not in pat_up:
        base = name_up.split("-")[0]
        return base == pat_up
    return False


def _auto_detect_type(raw_id: str) -> list[str]:
    """Given bare ID like '8.3.2', return candidates with type prefix."""
    parts = raw_id.split(".")
    if len(parts) == 3:
        return [f"S{raw_id}"]
    elif len(parts) == 2:
        return [f"L{raw_id}", f"C{raw_id}"]
    elif len(parts) == 1:
        return [f"C{raw_id}"]
    return [raw_id]


def _parse_query(query: str) -> tuple[str, str]:
    """Split query into (switch_name, port_filter).

    Supports: 'C1.15 20/2', 'C1.15  20/2', 'S5.3.1 1/1'.
    If no port given, port_filter is empty string.
    """
    q = query.strip()
    # Match: switch_name <space(s)> port (digits/slash)
    m = re.match(r'^(\S+)\s+(\d+/\d+)$', q)
    if m:
        return m.group(1), m.group(2)
    return q, ""


def _search(connections: list[dict], query: str) -> list[dict]:
    q = query.strip()
    if not q:
        return []

    switch_q, port_filter = _parse_query(q)

    # If it looks like a bare ID (digits and dots), auto-detect type
    has_prefix = bool(re.match(r'^[SCLNscln]', switch_q)) and not switch_q[0].isdigit()
    if has_prefix:
        candidates = [switch_q.upper()]
    elif re.match(r'^\d+[\.\d]*$', switch_q):
        candidates = [c.upper() for c in _auto_detect_type(switch_q)]
    else:
        candidates = [switch_q.upper()]

    # Also support substring match on switch names for partial queries
    results = []
    seen = set()
    sq_upper = switch_q.upper()

    for conn in connections:
        conn_id = id(conn)
        if conn_id in seen:
            continue

        src_up = conn["src_name"].upper()
        dest_up = conn["dest_name"].upper()

        matched = False
        matched_side = None  # track which side matched for port filtering
        for cand in candidates:
            if _switch_matches(src_up, cand):
                matched = True
                matched_side = "src"
                break
            if _switch_matches(dest_up, cand):
                matched = True
                matched_side = "dest"
                break

        if not matched:
            # Substring match — but for short queries like "L10", require
            # word boundary (dot, dash, end of string) after the match
            if len(sq_upper) <= 4:
                pat = re.compile(re.escape(sq_upper) + r'(?=[\.\-]|$)')
                if pat.search(src_up):
                    matched = True
                    matched_side = "src"
                elif pat.search(dest_up):
                    matched = True
                    matched_side = "dest"
            else:
                if sq_upper in src_up:
                    matched = True
                    matched_side = "src"
                elif sq_upper in dest_up:
                    matched = True
                    matched_side = "dest"

        if not matched:
            # Check fabric ID
            if conn.get("fabric_id") and sq_upper in conn["fabric_id"].upper():
                matched = True

        # Apply port filter if provided
        if matched and port_filter:
            port_ok = False
            if matched_side == "src" and conn["src_port"] == port_filter:
                port_ok = True
            elif matched_side == "dest" and conn["dest_port"] == port_filter:
                port_ok = True
            # Also check both sides if match was via fabric ID
            elif matched_side is None:
                if conn["src_port"] == port_filter or conn["dest_port"] == port_filter:
                    port_ok = True
            if not port_ok:
                matched = False

        if matched:
            seen.add(conn_id)
            results.append(conn)

    return results


# ════════════════════════════════════════════════════════════════════
#  DISPLAY
# ════════════════════════════════════════════════════════════════════

def _tier_label(conn: dict) -> str:
    """Short label like 'Spine→Core' or 'Leaf→Node'."""
    s = conn.get("src_type", "")
    d = conn.get("dest_type", "")
    if s and d:
        return f"{s}→{d}"
    return ""


def _cab_label(conn: dict, side: str) -> str:
    """Return 'Cab XX' string for a connection side, or empty."""
    cab = conn.get(f"{side}_cab", "")
    if cab:
        return f"Cab{cab}"
    return ""


def _print_result(conn: dict, idx: int):
    num = f"{BOLD}{idx:>3}.{RESET}"
    tier = f"{DIM}{_tier_label(conn)}{RESET}"
    line = (f"{CYAN}{conn['src_name']}{RESET} {conn['src_port']} "
            f"{DIM}→{RESET} "
            f"{YELLOW}{conn['dest_name']}{RESET} {conn['dest_port']}")
    print(f"  {num} {line}  {tier}")


def _short_fabric(fabric_id: str) -> str:
    """us-central-07a-fab52 → fab52"""
    if not fabric_id:
        return ""
    return fabric_id.rsplit("-", 1)[-1]


def _print_detail(conn: dict):
    tier = _tier_label(conn)
    dh = conn.get("data_hall", "")

    # Line 1: connection + tier
    line = (f"{CYAN}{BOLD}{conn['src_name']}{RESET} {conn['src_port']} "
            f"{DIM}→{RESET} "
            f"{YELLOW}{BOLD}{conn['dest_name']}{RESET} {conn['dest_port']}"
            f"  {DIM}{tier}{RESET}")
    print(f"\n  {line}")

    # Line 2: cable metadata
    meta = []
    if conn.get("status"):
        meta.append(conn["status"])
    if conn.get("cable_type"):
        meta.append(conn["cable_type"])
    if conn.get("cable_length") and conn["cable_length"] != "??":
        meta.append(conn["cable_length"])
    if conn.get("optic_type"):
        meta.append(conn["optic_type"])
    fab = _short_fabric(conn.get("fabric_id", ""))
    if fab:
        meta.append(fab)
    if meta:
        print(f"  {DIM}{' · '.join(meta)}{RESET}")

    # Line 3: racks + cabs
    rack_a = _extract_rack(conn["src_name"], dh)
    rack_b = _extract_rack(conn["dest_name"], dh)
    src_cab = _cab_label(conn, "src")
    dest_cab = _cab_label(conn, "dest")

    def _loc(rack, name, cab):
        parts = [f"{BOLD}R{rack}{RESET}"] if rack else []
        elev = _ELEVATIONS.get(name)
        if elev:
            parts.append(f"RU{elev['ru']}")
        if cab:
            parts.append(cab)
        label = " ".join(parts) if parts else ""
        return f"{label} {DIM}({name}){RESET}" if label else f"{DIM}({name}){RESET}"

    if rack_a or rack_b or src_cab or dest_cab:
        loc_a = _loc(rack_a, conn["src_name"], src_cab)
        loc_b = _loc(rack_b, conn["dest_name"], dest_cab)
        print(f"  {loc_a} → {loc_b}")


def _pick_layout(conn: dict) -> str:
    """Return primary DH layout key using elevation data."""
    halls = _map_halls(conn)
    return halls[0]


def _map_halls(conn: dict) -> list[str]:
    """Return list of hall keys to render maps for, using elevation data for truth."""
    # Use elevation data to determine actual DH for each side
    src_elev = _ELEVATIONS.get(conn["src_name"].upper())
    dest_elev = _ELEVATIONS.get(conn["dest_name"].upper())
    src_dh = src_elev["dh"] if src_elev and src_elev.get("dh") else None
    dest_dh = dest_elev["dh"] if dest_elev and dest_elev.get("dh") else None

    halls = set()
    if src_dh:
        halls.add(src_dh)
    else:
        # No elevation data — check switch name suffix (e.g. L20.1.1-DH2)
        name = conn.get("src_name", "")
        if "DH2" in name:
            halls.add("DH2")
        elif "DH1" in name:
            halls.add("DH1")
    if dest_dh:
        halls.add(dest_dh)
    else:
        name = conn.get("dest_name", "")
        if "DH2" in name:
            halls.add("DH2")
        elif "DH1" in name:
            halls.add("DH1")

    if not halls:
        dh = conn.get("data_hall", "")
        if dh in ("DH1", "DH2"):
            return [dh]
        return ["DH1"]

    return sorted(halls)


def _draw_elevation(conn: dict, searched: str = ""):
    """Draw a full rack elevation for both sides with actual RU positions."""
    src_name = conn["src_name"].upper()
    dest_name = conn["dest_name"].upper()

    src_elev = _ELEVATIONS.get(src_name)
    dest_elev = _ELEVATIONS.get(dest_name)

    if not src_elev and not dest_elev:
        print(f"  {DIM}No elevation data for this connection{RESET}")
        return

    def _rack_switch_map(rack_num: int) -> dict[int, dict]:
        """Get all switches in a rack as {ru: {name, sku}}."""
        by_ru: dict[int, dict] = {}
        for sw_name, info in _ELEVATIONS.items():
            if info["rack"] == rack_num:
                by_ru[info["ru"]] = {"name": sw_name, "sku": info.get("sku", "")}
        return by_ru

    highlights = {src_name, dest_name}

    def _render_rack(rack_num: int, cab: str, side_label: str) -> list[str]:
        """Render full rack with all RU positions, top to bottom."""
        by_ru = _rack_switch_map(rack_num)
        if not by_ru:
            return [f"  {DIM}(no elevation data for R{rack_num}){RESET}"]

        max_ru = max(by_ru.keys())
        min_ru = min(by_ru.keys())

        lines = [f"  {BOLD}Rack {rack_num}{RESET}  {DIM}{side_label}{RESET}"]
        lines.append(f"  {DIM}┌{'─' * 22}┐{RESET}")

        for ru in range(max_ru, min_ru - 1, -1):
            sw = by_ru.get(ru)
            if sw:
                name = sw["name"]
                if name in highlights:
                    lines.append(f"  {DIM}│{RESET} {BOLD}{ru:>2}{RESET}  {CYAN}{BOLD}{name}{RESET} {CYAN}{BOLD}◄{RESET}")
                else:
                    lines.append(f"  {DIM}│{RESET} {DIM}{ru:>2}{RESET}  {DIM}{name}{RESET}")
            else:
                lines.append(f"  {DIM}│ {ru:>2}  ·{RESET}")

        lines.append(f"  {DIM}└{'─' * 22}┘{RESET}")
        return lines

    print()
    sides = []
    no_elev_note = []
    if src_elev:
        src_cab = conn.get("src_cab", "")
        sides.append(_render_rack(src_elev["rack"], src_cab, conn["src_name"]))
    else:
        cab = conn.get("src_cab", "")
        no_elev_note.append(f"{conn['src_name']}{f' (Cab{cab})' if cab else ''}")
    if dest_elev:
        dest_cab = conn.get("dest_cab", "")
        if not src_elev or dest_elev["rack"] != src_elev["rack"]:
            sides.append(_render_rack(dest_elev["rack"], dest_cab, conn["dest_name"]))
    else:
        cab = conn.get("dest_cab", "")
        no_elev_note.append(f"{conn['dest_name']}{f' (Cab{cab})' if cab else ''}")

    _ansi_re = re.compile(r'\033\[[0-9;]*m')

    if len(sides) == 2:
        max_len = max(len(sides[0]), len(sides[1]))
        for s in sides:
            while len(s) < max_len:
                s.append("")
        max_w = max(len(_ansi_re.sub('', line)) for line in sides[0])
        pad = max(max_w + 4, 30)
        for l_line, r_line in zip(sides[0], sides[1]):
            visible_len = len(_ansi_re.sub('', l_line))
            spacing = pad - visible_len
            print(f"{l_line}{' ' * spacing}{r_line}")
    elif sides:
        for line in sides[0]:
            print(line)
    if no_elev_note:
        for name in no_elev_note:
            print(f"  {DIM}{name} — no rack elevation data (leaf switches not in ELEV tabs){RESET}")
    print()


def _detail_options_hint(has_map: bool, has_elev: bool) -> str:
    """Build the options hint string."""
    opts = []
    if has_map:
        opts.append(f"{DIM}[m]{RESET} map")
    if has_elev:
        opts.append(f"{DIM}[e]{RESET} elevation")
    opts.append(f"{DIM}[t]{RESET} tips")
    opts.append(f"{DIM}[Enter]{RESET} back")
    return " ".join(opts)


def _detail_prompt(conn: dict):
    """Handle [m]/[t]/[e] sub-prompt after compact detail."""
    dh = conn.get("data_hall", "")
    rack_a = _extract_rack(conn["src_name"], dh)
    rack_b = _extract_rack(conn["dest_name"], dh)
    has_map = bool(rack_a or rack_b)
    has_elev = bool(
        _ELEVATIONS.get(conn["src_name"].upper())
        or _ELEVATIONS.get(conn["dest_name"].upper())
    )

    print(f"  {_detail_options_hint(has_map, has_elev)}")

    while True:
        sel = _prompt()
        if not sel:
            return  # back to search — caller handles newline
        sl = sel.lower()
        if sl in ("q", "quit", "exit"):
            raise SystemExit(0)
        if sl == "m" and has_map:
            halls = _map_halls(conn)
            src_elev = _ELEVATIONS.get(conn["src_name"].upper())
            dest_elev = _ELEVATIONS.get(conn["dest_name"].upper())
            src_dh = src_elev["dh"] if src_elev and src_elev.get("dh") else ("DH2" if "DH2" in conn["src_name"] else "DH1" if "DH1" in conn["src_name"] else None)
            dest_dh = dest_elev["dh"] if dest_elev and dest_elev.get("dh") else ("DH2" if "DH2" in conn["dest_name"] else "DH1" if "DH1" in conn["dest_name"] else None)
            for lk in halls:
                ha = rack_a if (src_dh == lk or src_dh is None) else None
                hb = rack_b if (dest_dh == lk or dest_dh is None) else None
                la = conn["src_name"] if ha else ""
                lb = conn["dest_name"] if hb else ""
                layout = _LAYOUTS.get(lk, _LAYOUTS["DH1"])
                _draw_map(layout, highlight_a=ha, highlight_b=hb, label_a=la, label_b=lb)
            print(f"  {_detail_options_hint(has_map, has_elev)}")
        elif sl == "e" and has_elev:
            _draw_elevation(conn)
            print(f"  {_detail_options_hint(has_map, has_elev)}")
        elif sl == "t":
            tips = _get_tips(conn)
            if tips:
                print(f"\n  {BOLD}Steps:{RESET}")
                for i, (label, cmd) in enumerate(tips, 1):
                    print(f"    {DIM}{i}.{RESET} {label}")
                    print(f"       {CYAN}{cmd}{RESET}")
                print()
            print(f"  {_detail_options_hint(has_map, has_elev)}")
        else:
            return  # unrecognized input — back to search


# ════════════════════════════════════════════════════════════════════
#  INTERACTIVE LOOP
# ════════════════════════════════════════════════════════════════════

def _flush_stdin():
    """Flush any buffered keystrokes so they don't leak into the next prompt."""
    try:
        import termios
        termios.tcflush(sys.stdin, termios.TCIFLUSH)
    except (ImportError, termios.error, ValueError):
        pass


def _prompt(hint: str = "> ") -> str:
    try:
        return input(f"  {hint}").strip()
    except (EOFError, KeyboardInterrupt):
        return "q"


def _clear():
    os.system("clear" if os.name != "nt" else "cls")


def _run():
    global _ELEVATIONS
    _clear()
    print(f"\n  {DIM}Loading EVI01 data...{RESET}", end="", flush=True)
    connections, _ELEVATIONS = _load_data()
    _clear()

    print(f"\n  {BOLD}IB Lookup{RESET}  {DIM}EVI01 · {len(connections):,} connections{RESET}")
    print(f"  {DIM}Search any switch — example: S5.3.1, L10, C1.15 20/2{RESET}\n")

    while True:
        query = _prompt(f"{BOLD}Search:{RESET} ")

        if query in ("q", "quit", "exit"):
            break
        if query == "?":
            print(f"\n  {BOLD}Examples:{RESET}")
            print(f"    S5.3.1       {DIM}— spine switch{RESET}")
            print(f"    L10          {DIM}— all leaf switches in cab 10{RESET}")
            print(f"    C1.15 20/2   {DIM}— core switch + specific port{RESET}")
            print(f"    8.3.2        {DIM}— auto-detects type{RESET}")
            print(f"    q            {DIM}— quit{RESET}\n")
            continue
        if not query:
            continue

        results = _search(connections, query)
        if not results:
            print(f"  {DIM}No matches for '{query}'{RESET}\n")
            continue

        print(f"\n  {BOLD}{query}{RESET} — {len(results)} match{'es' if len(results) != 1 else ''}\n")
        for idx, conn in enumerate(results, 1):
            _print_result(conn, idx)

        if len(results) == 1:
            _print_detail(results[0])
            _detail_prompt(results[0])
            _flush_stdin()
            continue

        print(f"\n  {DIM}Enter a number for details, or just search again{RESET}")
        sel = _prompt()
        if sel in ("q", "quit", "exit"):
            break
        if sel.isdigit():
            n = int(sel) - 1
            if 0 <= n < len(results):
                _print_detail(results[n])
                _detail_prompt(results[n])
                _flush_stdin()
            else:
                print(f"  {DIM}Pick 1–{len(results)}{RESET}")


# ════════════════════════════════════════════════════════════════════
#  DROP ZONE — inline file loader when xlsx is missing
# ════════════════════════════════════════════════════════════════════

_AMBER = "\033[38;5;208m"
_BORDER = "\033[38;5;238m"
_LGRAY = "\033[38;5;245m"
_GRAY = "\033[38;5;240m"
_W = 54

def _pad(text):
    left = (_W - len(text)) // 2
    right = _W - len(text) - left
    return " " * left + text + " " * right

def _box_line(content):
    print(f"  {_BORDER}│{RESET}{content}{_BORDER}│{RESET}")

def _box_blank():
    _box_line(" " * _W)

def _box_top():
    print(f"  {_BORDER}┌{'─' * _W}┐{RESET}")

def _box_bottom():
    print(f"  {_BORDER}└{'─' * _W}┘{RESET}")

def _prompt_drop_zone(sketch_url, sketch_xlsx, project_dir):
    """Show visual drop zone, accept file, copy it, then continue."""
    import shutil

    print("\033[2J\033[H")  # clear screen
    print()
    print(f"  {BOLD}{WHITE}IB Burndown{RESET}  {DIM}{_LGRAY}Sketch Loader{RESET}")
    print(f"  {DIM}{_GRAY}EVI01 · US-CENTRAL-07A · Elk Grove{RESET}")
    print()
    _box_top()
    _box_blank()
    _box_blank()
    _box_line(f"{_AMBER}{_pad('↑')}{RESET}")
    _box_blank()
    _box_line(f"{WHITE}{BOLD}{_pad('Drop .xlsx here')}{RESET}")
    _box_blank()
    _box_line(f"{DIM}{_LGRAY}{_pad('drag IB Sketch from Finder into terminal')}{RESET}")
    _box_line(f"{DIM}{_GRAY}{_pad('then press Enter')}{RESET}")
    _box_blank()
    _box_blank()
    _box_bottom()
    print()
    print(f"  {DIM}{_LGRAY}Source sheet:{RESET}")
    print(f"  {_AMBER}{sketch_url}{RESET}")
    print(f"  {DIM}{_GRAY}File → Download → Microsoft Excel (.xlsx){RESET}")
    print()

    filepath = input(f"  {_AMBER}→{RESET} ").strip().strip("'\"").rstrip()

    if not os.path.isfile(filepath):
        print(f"\n  {RED}{BOLD}✗{RESET}  {RED}File not found{RESET}")
        sys.exit(1)

    if not filepath.lower().endswith((".xlsx", ".xls")):
        ext = os.path.splitext(filepath)[1]
        print(f"\n  {RED}{BOLD}✗{RESET}  {RED}Expected .xlsx, got {ext}{RESET}")
        sys.exit(1)

    shutil.copy2(filepath, sketch_xlsx)
    size = os.path.getsize(sketch_xlsx)
    size_h = f"{size / 1024:.0f}KB" if size < 1048576 else f"{size / 1048576:.1f}MB"
    basename = os.path.basename(filepath)

    print("\033[2J\033[H")  # clear screen
    print()
    print(f"  {BOLD}{WHITE}IB Burndown{RESET}  {DIM}{_LGRAY}Sketch Loader{RESET}")
    print(f"  {DIM}{_GRAY}EVI01 · US-CENTRAL-07A · Elk Grove{RESET}")
    print()
    _box_top()
    _box_blank()
    _box_blank()
    _box_line(f"{GREEN}{BOLD}{_pad('✓')}{RESET}")
    _box_blank()
    _box_line(f"{GREEN}{BOLD}{_pad(basename[:_W])}{RESET}")
    _box_blank()
    _box_line(f"{_LGRAY}{_pad(f'{size_h} → EVI01 - IB Sketch.xlsx')}{RESET}")
    _box_line(f"{_LGRAY}{_pad('Launching...')}{RESET}")
    _box_blank()
    _box_blank()
    _box_bottom()
    print()

    import time
    time.sleep(1)


# ════════════════════════════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        prog="ib-lookup",
        description="EVI01 IB Lookup — search switch connections + map + tips",
    )
    parser.add_argument("query", nargs="?", default=None,
                        help="Switch name to search (e.g. L10, S5.3, C1.4)")
    args = parser.parse_args()

    _SKETCH_URL = "https://docs.google.com/spreadsheets/d/1U132alRVDtcrVd5kW4v534U3ME7wRZ5g3kHQMZP2LaM/edit?gid=1992819001#gid=1992819001"

    if not os.path.isfile(_SKETCH_XLSX):
        _prompt_drop_zone(_SKETCH_URL, _SKETCH_XLSX, _DIR)

    print(f"  {GREEN}File found:{RESET} {os.path.basename(_SKETCH_XLSX)}")

    if args.query:
        global _ELEVATIONS
        connections, _ELEVATIONS = _load_data()
        results = _search(connections, args.query)
        if not results:
            print(f"  {DIM}No matches for '{args.query}'{RESET}")
            return
        print(f"\n  {BOLD}{args.query}{RESET} — {len(results)} matches\n")
        for idx, conn in enumerate(results, 1):
            _print_result(conn, idx)
        if len(results) == 1:
            _print_detail(results[0])
        print()
        return

    _run()


if __name__ == "__main__":
    main()
