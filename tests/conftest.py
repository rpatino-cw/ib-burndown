"""Shared fixtures for ib-lookup tests."""

import json
import os
import pytest
import tempfile

import openpyxl


@pytest.fixture
def tmp_dir(tmp_path):
    """Provide a temp directory for test files."""
    return tmp_path


@pytest.fixture
def sample_ib_xlsx(tmp_path):
    """Create a minimal IB Sketch xlsx with Pull Schedule + ELEV tabs."""
    path = tmp_path / "Test-Site - IB Sketch.xlsx"
    wb = openpyxl.Workbook()

    # Pull Schedule tab
    ws = wb.active
    ws.title = "Core Group 1 Pull Schedule"
    ws.append(["Status", "Source", "Source Port", "Destination", "Destination Port",
               "Cable Type", "Cable Length", "Optic Type", "Fabric ID"])
    ws.append(["Active", "S1.1.1", "1/1", "C1.1", "9/1", "MTP", "5m", "Twin Port OSFP", "test-site-fab01"])
    ws.append(["Active", "S1.1.1", "1/2", "C1.2", "9/1", "MTP", "5m", "Twin Port OSFP", "test-site-fab01"])
    ws.append(["Cable Not Run", "S1.1.1", "2/1", "C1.3", "9/1", "MTP", "10m", "Twin Port OSFP", "test-site-fab01"])

    # DH1 Pull Schedule
    ws2 = wb.create_sheet("DH1 Pull Schedule")
    ws2.append(["Status", "Source", "Source Port", "Destination", "Destination Port",
                "Cable Type", "Cable Length", "Optic Type", "Fabric ID"])
    ws2.append(["Active", "L10.1.1-DH1", "25/1", "S1.1.1", "17/1", "MTP", "3m", "Twin Port OSFP", "test-site-fab01"])

    # DH2 Pull Schedule
    ws3 = wb.create_sheet("DH2 Pull Schedule")
    ws3.append(["Status", "Source", "Source Port", "Destination", "Destination Port",
                "Cable Type", "Cable Length", "Optic Type", "Fabric ID"])
    ws3.append(["Active", "L20.1.1-DH2", "25/1", "S2.1.1", "21/1", "MTP", "3m", "Twin Port OSFP", "test-site-fab01"])

    # DH1 ELEV tab
    ws_elev = wb.create_sheet("DH1 ELEV")
    ws_elev.append([None, None, None])  # row 1
    ws_elev.append([None, None, "DH1 Row 5"])  # row 2 — label
    ws_elev.append([None, None, None, 41, None, 43])  # row 3 — rack numbers
    ws_elev.append([38, None, None, "S1.1.1", "MQM9790-NS2F", "C1.1", "MQM9790"])  # RU 38
    ws_elev.append([37, None, None, "S1.1.2", "MQM9790-NS2F", "C1.2", "MQM9790"])  # RU 37

    # DH2 ELEV tab
    ws_elev2 = wb.create_sheet("DH2 ELEV")
    ws_elev2.append([None, None, None])
    ws_elev2.append([None, None, "DH2 Row 1"])
    ws_elev2.append([None, None, None, 10, None, 20])
    ws_elev2.append([38, None, None, "S2.1.1", "MQM9790", "L20.1.1-DH2", "IB Leaf"])

    wb.save(path)
    return str(path)


@pytest.fixture
def sample_trad_json(tmp_path):
    """Create a minimal trad-data.json for testing."""
    data = {
        "site": "us-test-01",
        "data_halls": ["dh1", "dh2"],
        "devices": [
            {"dns": "dh2-t0-d1-01-r001-us-test-01", "loc": "dh2:1:46", "dh": "dh2",
             "rack": 1, "ru": 46, "model": "SN5610", "role": "t0", "mgmt_ip": "10.0.0.1"},
            {"dns": "dh2-infra-d1-01-r130-us-test-01", "loc": "dh2:130:44", "dh": "dh2",
             "rack": 130, "ru": 44, "model": "SN3700", "role": "infra", "mgmt_ip": ""},
            {"dns": "dh2-net-d1-01-r001-us-test-01", "loc": "dh2:1:44", "dh": "dh2",
             "rack": 1, "ru": 44, "model": "SN2201", "role": "net", "mgmt_ip": ""},
        ],
        "connections": [
            {"status": "Active", "a_dns": "dh2-t0-d1-01-r001-us-test-01", "a_port": "swp1",
             "a_rack": 1, "a_ru": 46, "a_dh": "dh2",
             "z_dns": "dh2-infra-d1-01-r130-us-test-01", "z_port": "swp31",
             "z_rack": 130, "z_ru": 44, "z_dh": "dh2", "cable": "LC-TO-LC SMF"},
            {"status": "Active", "a_dns": "dh2-t0-d1-01-r001-us-test-01", "a_port": "swp2",
             "a_rack": 1, "a_ru": 46, "a_dh": "dh2",
             "z_dns": "dh2-net-d1-01-r001-us-test-01", "z_port": "swp1",
             "z_rack": 1, "z_ru": 44, "z_dh": "dh2", "cable": "DAC"},
        ],
        "conn_index": {
            "dh2-t0-d1-01-r001-us-test-01": [0, 1],
            "dh2-infra-d1-01-r130-us-test-01": [0],
            "dh2-net-d1-01-r001-us-test-01": [1],
        },
        "rack_inventory": {
            "1": [
                {"dns": "dh2-t0-d1-01-r001-us-test-01", "ru": 46, "model": "SN5610", "role": "t0"},
                {"dns": "dh2-net-d1-01-r001-us-test-01", "ru": 44, "model": "SN2201", "role": "net"},
            ],
            "130": [
                {"dns": "dh2-infra-d1-01-r130-us-test-01", "ru": 44, "model": "SN3700", "role": "infra"},
            ],
        },
    }
    path = tmp_path / "trad-data.json"
    with open(path, "w") as f:
        json.dump(data, f)
    return str(path)


@pytest.fixture
def sample_overhead_csv(tmp_path):
    """Create a minimal overhead CSV for layout import testing."""
    csv_content = """,,,,,,,,,,,
,US-TEST01 DH1,,,,,,,,,,
,,,,,,,,,,,,
,,1,2,3,4,5,6,7,8,9,10,
,,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,
,,10,9,8,7,6,5,4,3,2,1,
,,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,HD-B2,
"""
    path = tmp_path / "overhead.csv"
    path.write_text(csv_content)
    return str(path)


@pytest.fixture
def sample_layouts():
    """Return a sample layouts dict."""
    return {
        "TEST.DH1": {
            "racks_per_row": 10,
            "columns": [
                {"label": "Left", "start": 1, "num_rows": 5, "racks_per_row": 10},
                {"label": "Right", "start": 51, "num_rows": 5, "racks_per_row": 10},
            ],
            "serpentine": True,
        }
    }
