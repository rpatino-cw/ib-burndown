"""Traditional networking parser — extract devices, connections, and layout from MASTER cutsheet.

Reads SITE-HOSTS, CUTSHEET, and OVERHEAD tabs from any site's MASTER xlsx.
Site-agnostic — auto-detects DH, devices, and connections.
"""

from __future__ import annotations

import json
import os
import re
from collections import defaultdict

import openpyxl


def _parse_loc(loc_str: str) -> tuple[str, int, int]:
    """Parse 'dh2:130:44' into (dh, rack, ru)."""
    parts = str(loc_str).split(":")
    dh = parts[0] if parts else ""
    rack = 0
    ru = 0
    if len(parts) >= 2:
        try:
            rack = int(parts[1])
        except ValueError:
            pass
    if len(parts) >= 3:
        try:
            ru = int(parts[2])
        except ValueError:
            pass
    return dh, rack, ru


def _find_header_cols(ws, target_headers: dict[str, list[str]]) -> dict[str, int]:
    """Find column indices by matching header names (case-insensitive).

    target_headers: {field_name: [possible_header_strings]}
    Returns: {field_name: col_index}
    """
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for i, h in enumerate(row):
            if h is None:
                continue
            hl = str(h).strip().lower()
            for field, patterns in target_headers.items():
                if field in col_map:
                    continue
                for pat in patterns:
                    if pat.lower() in hl:
                        col_map[field] = i
                        break
        break
    return col_map


def detect_trad_site(wb) -> tuple[str, list[str]]:
    """Auto-detect site and DHs from SITE-HOSTS locations."""
    dh_set = set()
    site = ""

    if "SITE-HOSTS" not in wb.sheetnames:
        return "", []

    ws = wb["SITE-HOSTS"]
    col_map = _find_header_cols(ws, {"loc": ["location", "loc"], "dns": ["dns", "hostname"]})
    loc_col = col_map.get("loc")
    dns_col = col_map.get("dns")

    if loc_col is None:
        return "", []

    for row in ws.iter_rows(min_row=2, values_only=True):
        loc = str(row[loc_col]) if loc_col < len(row) and row[loc_col] else ""
        dh_match = re.match(r'^(dh\d+)', loc, re.IGNORECASE)
        if dh_match:
            dh_set.add(dh_match.group(1).lower())

        # Extract site from DNS: dh2-xxx-...-us-central-07a → us-central-07a
        if not site and dns_col and dns_col < len(row) and row[dns_col]:
            dns = str(row[dns_col])
            site_m = re.search(r'((?:us|gb|se|no|de|fr)-[\w]+-[\w]+)$', dns, re.IGNORECASE)
            if site_m:
                site = site_m.group(1)

    return site, sorted(dh_set)


def extract_devices(wb, dh_filter: str = "") -> list[dict]:
    """Extract networking devices from SITE-HOSTS.

    If dh_filter is set (e.g. 'dh2'), only return devices in that DH.
    """
    if "SITE-HOSTS" not in wb.sheetnames:
        return []

    ws = wb["SITE-HOSTS"]
    col_map = _find_header_cols(ws, {
        "loc": ["location", "loc"],
        "dns": ["dns", "hostname"],
        "model": ["model", "netbox model"],
        "role": ["role"],
        "mgmt_ip": ["mgmt", "management ip"],
    })

    devices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def _g(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        loc = _g("loc")
        dns = _g("dns")
        model = _g("model")
        role = _g("role")
        mgmt_ip = _g("mgmt_ip")

        if not dns:
            continue

        dh, rack, ru = _parse_loc(loc)

        if dh_filter and dh.lower() != dh_filter.lower():
            continue

        # Skip IB devices — this is traditional only
        if "ib-" in dns.lower() or "QM97" in model:
            continue

        devices.append({
            "dns": dns,
            "loc": loc,
            "dh": dh,
            "rack": rack,
            "ru": ru,
            "model": model,
            "role": role,
            "mgmt_ip": mgmt_ip,
        })

    return devices


def extract_all_devices(wb, dh_filter: str = "") -> list[dict]:
    """Extract ALL devices from SITE-HOSTS for full rack elevations."""
    if "SITE-HOSTS" not in wb.sheetnames:
        return []

    ws = wb["SITE-HOSTS"]
    col_map = _find_header_cols(ws, {
        "loc": ["location", "loc"],
        "dns": ["dns", "hostname"],
        "model": ["model", "netbox model"],
        "role": ["role"],
    })

    devices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def _g(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        loc = _g("loc")
        dns = _g("dns")
        if not dns:
            continue

        dh, rack, ru = _parse_loc(loc)
        if not rack:
            continue
        if dh_filter and dh.lower() != dh_filter.lower():
            continue

        devices.append({
            "dns": dns,
            "dh": dh,
            "rack": rack,
            "ru": ru,
            "model": _g("model"),
            "role": _g("role"),
        })

    return devices


def extract_connections(wb, dh_filter: str = "") -> list[dict]:
    """Extract traditional networking connections from CUTSHEET."""
    if "CUTSHEET" not in wb.sheetnames:
        return []

    ws = wb["CUTSHEET"]
    col_map = _find_header_cols(ws, {
        "status": ["status"],
        "a_loc": ["a-side-loc", "a side loc", "a_loc", "a loc"],
        "a_dns": ["a-side-dns", "a side dns", "a_dns", "a dns"],
        "a_model": ["a-side-model", "a model"],
        "a_port": ["a-port", "a port", "a_port"],
        "z_loc": ["z-side-loc", "z side loc", "z_loc", "z loc"],
        "z_dns": ["z-side-dns", "z side dns", "z_dns", "z dns"],
        "z_model": ["z-side-model", "z model"],
        "z_port": ["z-port", "z port", "z_port"],
        "cable": ["cable"],
    })

    conns = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def _g(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        a_loc = _g("a_loc")
        z_loc = _g("z_loc")
        a_port = _g("a_port")
        z_port = _g("z_port")

        # Traditional ports only (swp/eth)
        has_trad = ("swp" in a_port or "eth" in a_port
                    or "swp" in z_port or "eth" in z_port)
        if not has_trad:
            continue

        a_dh, a_rack, a_ru = _parse_loc(a_loc)
        z_dh, z_rack, z_ru = _parse_loc(z_loc)

        if dh_filter:
            if a_dh.lower() != dh_filter.lower() and z_dh.lower() != dh_filter.lower():
                continue

        conns.append({
            "status": _g("status"),
            "a_dns": _g("a_dns"),
            "a_port": a_port,
            "a_rack": a_rack,
            "a_ru": a_ru,
            "a_dh": a_dh,
            "z_dns": _g("z_dns"),
            "z_port": z_port,
            "z_rack": z_rack,
            "z_ru": z_ru,
            "z_dh": z_dh,
            "cable": _g("cable"),
        })

    return conns


def export_trad_data(xlsx_path: str, output_path: str | None = None,
                     dh_filter: str = "") -> dict:
    """Parse MASTER cutsheet and export trad-data.json.

    Returns the data dict.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    site, dh_list = detect_trad_site(wb)
    devices = extract_devices(wb, dh_filter)
    all_devs = extract_all_devices(wb, dh_filter)
    connections = extract_connections(wb, dh_filter)
    wb.close()

    # Build connection index
    conn_index = defaultdict(list)
    for i, c in enumerate(connections):
        if c["a_dns"]:
            conn_index[c["a_dns"]].append(i)
        if c["z_dns"] and c["z_dns"] != c["a_dns"]:
            conn_index[c["z_dns"]].append(i)

    # Build rack inventory
    rack_inventory = defaultdict(list)
    for d in all_devs:
        rack_inventory[d["rack"]].append({
            "dns": d["dns"], "ru": d["ru"],
            "model": d["model"], "role": d["role"],
        })
    for rack in rack_inventory:
        rack_inventory[rack].sort(key=lambda x: -x["ru"])

    data = {
        "site": site,
        "data_halls": dh_list,
        "devices": devices,
        "connections": connections,
        "conn_index": dict(conn_index),
        "rack_inventory": {str(k): v for k, v in rack_inventory.items()},
    }

    if output_path:
        with open(output_path, "w") as f:
            json.dump(data, f)

    return data
