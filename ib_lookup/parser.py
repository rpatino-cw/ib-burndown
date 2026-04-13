"""Excel parsing — connections + elevations from any IB Sketch xlsx."""

from __future__ import annotations

import re
from datetime import datetime

import openpyxl

_TYPE_PREFIX = {"Spine": "S", "Core": "C", "Leaf": "L", "Node": "N"}
_PREFIX_TYPE = {"S": "Spine", "C": "Core", "L": "Leaf", "N": "Node"}


def normalize_port(value) -> str:
    """Convert datetime-corrupted port values back to 'port/lane'."""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, (int, float)):
        i = int(value)
        return str(i) if value == i else str(value)
    if value is None:
        return ""
    return str(value).strip()


def _normalize_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _normalize_cab(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        i = int(value)
        return str(i) if value == i else str(value)
    return str(value).strip()


def build_switch_name(type_str: str, switch_id: str, data_hall: str, cab: str = "") -> str:
    """Build canonical switch name: S1.1.1, L10.1.2-DH2, C1.4."""
    prefix = _TYPE_PREFIX.get(type_str, "")
    if type_str == "Leaf" and cab:
        name = f"{prefix}{cab}.{switch_id}"
        if data_hall:
            name += f"-{data_hall}"
    else:
        name = f"{prefix}{switch_id}"
    return name


def parse_switch_name(name: str) -> dict:
    """Parse 'C1.17', 'S1.5.2', 'L10.1.2-DH2' into {type, dh, cab, id}."""
    result = {"type": "", "dh": "", "cab": "", "id": ""}
    if not name:
        return result

    dh_m = re.search(r'-DH(\d+)$', name, re.IGNORECASE)
    base = name[:dh_m.start()] if dh_m else name
    result["dh"] = f"DH{dh_m.group(1)}" if dh_m else ""

    if base and base[0].upper() in _PREFIX_TYPE:
        result["type"] = _PREFIX_TYPE[base[0].upper()]
        base = base[1:]
    else:
        return result

    if result["type"] == "Leaf":
        m = re.match(r'^(\d+)\.(.+)$', base)
        if m:
            result["cab"] = m.group(1)
            result["id"] = m.group(2)
        else:
            result["id"] = base
    else:
        result["id"] = base

    return result


def detect_site(wb) -> tuple[str, list[str]]:
    """Auto-detect site name and data halls from Excel sheet names.

    Returns (site_name, list_of_dh_keys).
    """
    dh_set = set()
    site_hints = set()

    for tab_name in wb.sheetnames:
        # Find DH references
        for m in re.finditer(r'(DH\d+)', tab_name, re.IGNORECASE):
            dh_set.add(m.group(1).upper())

        # Look for site name patterns in ELEV/schedule tabs
        site_m = re.match(r'^([\w-]+)\.(DH\d+)', tab_name)
        if site_m:
            site_hints.add(site_m.group(1))

    site_name = next(iter(site_hints)) if site_hints else ""
    dh_list = sorted(dh_set) if dh_set else ["DH1"]

    return site_name, dh_list


def parse_connections(path: str) -> list[dict]:
    """Parse Pull Schedule tabs from IB Sketch into connections list.

    Works with any site — no hardcoded site names.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    connections: list[dict] = []

    for tab_name in wb.sheetnames:
        if "Pull Schedule" not in tab_name:
            continue

        # Detect DH from tab name dynamically
        dh_match = re.search(r'(DH\d+)', tab_name, re.IGNORECASE)
        tab_dh = dh_match.group(1).upper() if dh_match else ""

        ws = wb[tab_name]
        header_row = None
        col_map = {}

        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header_row = row
            break

        if not header_row:
            continue

        for i, h in enumerate(header_row):
            if h is None:
                continue
            hl = str(h).strip().lower()
            if hl == "status":
                col_map["status"] = i
            elif hl == "source":
                col_map["source"] = i
            elif hl == "source port":
                col_map["src_port"] = i
            elif hl == "destination":
                col_map["dest"] = i
            elif hl == "destination port":
                col_map["dest_port"] = i
            elif hl == "cable type":
                col_map["cable_type"] = i
            elif "optic" in hl and "type" in hl:
                col_map["optic_type"] = i
            elif hl == "cable length":
                col_map["cable_length"] = i
            elif hl == "fabric id":
                col_map["fabric_id"] = i

        if "source" not in col_map or "dest" not in col_map:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            def _g(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                val = row[idx]
                if isinstance(val, datetime):
                    return f"{val.month}/{val.day}"
                return str(val).strip()

            src_name = _g("source")
            dest_name = _g("dest")
            if not src_name or not dest_name:
                continue

            src_info = parse_switch_name(src_name)
            dest_info = parse_switch_name(dest_name)
            data_hall = src_info["dh"] or dest_info["dh"] or tab_dh

            connections.append({
                "data_hall": data_hall,
                "src_type": src_info["type"],
                "src_dh": src_info["dh"] or tab_dh,
                "src_cab": src_info["cab"],
                "src_id": src_info["id"],
                "src_port": _g("src_port"),
                "src_name": src_name,
                "dest_type": dest_info["type"],
                "dest_dh": dest_info["dh"] or tab_dh,
                "dest_cab": dest_info["cab"],
                "dest_id": dest_info["id"],
                "dest_port": _g("dest_port"),
                "dest_name": dest_name,
                "tab_ref": tab_name,
                "status": _g("status"),
                "cable_type": _g("cable_type"),
                "cable_length": _g("cable_length"),
                "optic_type": _g("optic_type"),
                "fabric_id": _g("fabric_id"),
            })

    wb.close()
    return connections


def parse_elevations(path: str) -> dict[str, dict]:
    """Parse ELEV tabs + Leaf Pull Schedule tabs.

    Returns lookup: SWITCH_NAME_UPPER → {rack, ru, sku, dh, row}.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    elevations: dict[str, dict] = {}

    for tab_name in wb.sheetnames:
        if "ELEV" not in tab_name:
            continue

        ws = wb[tab_name]
        rows_data = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows_data.append(row)
            if len(rows_data) >= 55:
                break

        if len(rows_data) < 4:
            continue

        # Row 2 (idx 1): DH + Row label
        row2 = rows_data[1]
        row_label = str(row2[2]).strip() if len(row2) > 2 and row2[2] else ""

        # Detect DH from tab name dynamically
        dh_match = re.search(r'(DH\d+)', tab_name, re.IGNORECASE)
        dh = dh_match.group(1).upper() if dh_match else ""

        # Row 3 (idx 2): rack numbers at cols 3, 5, 7, ...
        row3 = rows_data[2]
        rack_cols = {}
        for ci in range(3, len(row3), 2):
            val = row3[ci]
            if val is not None:
                try:
                    rack_cols[ci] = int(float(str(val)))
                except (ValueError, TypeError):
                    pass

        # Data rows (idx 3+): col 0 = RU, rack cols = switch names
        for row in rows_data[3:]:
            if not row or row[0] is None:
                continue
            try:
                ru = int(float(str(row[0])))
            except (ValueError, TypeError):
                continue

            for ci, rack_num in rack_cols.items():
                if ci >= len(row) or row[ci] is None:
                    continue
                switch_name = str(row[ci]).strip()
                if not switch_name or switch_name in ("Labels", "POWER", "Rack#"):
                    continue

                sku = ""
                if ci + 1 < len(row) and row[ci + 1] is not None:
                    sku = str(row[ci + 1]).strip()
                    if sku == "CW-SKU":
                        sku = ""

                elevations[switch_name.upper()] = {
                    "rack": rack_num,
                    "ru": ru,
                    "sku": sku,
                    "dh": dh,
                    "row": row_label,
                }

    # Parse Leaf Pull Schedule tabs
    leaf_tab_re = re.compile(r'(DH\d+)\s+Rack\s+(\d+)\s+Leaf Pull Schedule', re.IGNORECASE)
    for tab_name in wb.sheetnames:
        m = leaf_tab_re.match(tab_name)
        if not m:
            continue
        dh = m.group(1).upper()
        rack_num = int(m.group(2))

        ws = wb[tab_name]
        leaf_names = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1]:
                leaf_names.add(str(row[1]).strip().upper())

        sorted_leaves = sorted(leaf_names)
        count = len(sorted_leaves)
        for i, name in enumerate(sorted_leaves):
            if name not in elevations:
                elevations[name] = {
                    "rack": rack_num,
                    "ru": count - i,
                    "sku": "IB Leaf",
                    "dh": dh,
                    "row": "",
                }

    wb.close()
    return elevations
